from bs4 import BeautifulSoup as bs
from selenium import webdriver
import openpyxl
import datetime

def write_excel(raw_datas):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = '쇼핑몰명 데이터'
    sheet['A1'] = '다음 쇼핑몰명 현황'
    sheet['A3'] = '수집 일시'
    sheet['B3'] = datetime.datetime.now()
    sheet['A5'] = '위치'
    sheet['B5'] = '업체명'
    sheet['C5'] = '연결 URL'

    for i in range(len(raw_datas['icon'])):
        sheet.cell(row=i + 6, column=1).value = '아이콘_' + str(i+1)
        sheet.cell(row=i + 6, column=2).value = raw_datas['icon'][i][0]
        sheet.cell(row=i + 6, column=3).value = raw_datas['icon'][i][1]

    max = sheet.max_row + 1
    for i in range(len(raw_datas['text'])):
        for j in range(len(raw_datas['text'][i])):
            sheet.cell(row=max + j, column=1).value = '텍스트' + str(i+1) + '_' + str(j+1)
            sheet.cell(row=max + j, column=2).value = raw_datas['text'][i][j][0]
            sheet.cell(row=max + j, column=3).value = raw_datas['text'][i][j][1]
        max = sheet.max_row + 1

    max = sheet.max_row
    for i in range(len(raw_datas['mobile'])):
        for j in range(len(raw_datas['mobile'][i])):
            sheet.cell(row=max + j, column=1).value = '모바일' + str(i+1) + '_' + str(j+1)
            sheet.cell(row=max + j, column=2).value = raw_datas['mobile'][i][j][0]
            sheet.cell(row=max + j, column=3).value = raw_datas['mobile'][i][j][1]
        max = sheet.max_row + 1

    wb.save('result.xlsx')


def get_data(url):
    path = 'C:/chromedriver'
    driver = webdriver.Chrome(path)
    driver.get(url)

    if url.__contains__('//m.'):
        cl = driver.find_element_by_xpath('//*[@id="mArticle"]/nav[1]/div/div/ul/li[2]/a').click()

    elements = driver.find_elements_by_xpath('//*[contains(@id,"adfit_frame_id")]')
    origin_source = ''
    for i in range(len(elements)):
        if elements[i].get_attribute('outerHTML').__contains__('쇼핑'):
            driver.switch_to.frame(elements[i])
            origin_source = driver.page_source
            driver.switch_to.default_content()
            break

    driver.close()
    return origin_source

def parse_data(datas, case):
    soup = bs(datas, 'html.parser')

    if case == 'pc':
        pc_icon_link = soup.select('a.link_tab')
        pc_icon_name = soup.select('span.txt_tab')
        pc_text_lists = soup.select('ul.list_mall')

        pc_icon_datas = []
        for i, j in zip(pc_icon_name, pc_icon_link):
            pc_icon_datas.append([i.text, j.attrs['href']])
        print(pc_icon_datas)

        souplists = pc_text_lists[0].select('ul.list_inner')
        pc_text_datas = []
        for i in range(len(souplists)):
            soupdatas = souplists[i].select('li > a.link_mall')
            temp = []
            for j in soupdatas:
                temp.append([j.text, j.attrs['href']])
            pc_text_datas.append(temp)
        print(pc_text_datas)

        pc_datas = {
            'icon':pc_icon_datas,
            'text':pc_text_datas
        }
        return pc_datas

    elif case == 'mobile':
        mobile_lists = soup.select('div.inner_direct')
        datas = []
        for i in range(len(mobile_lists)):
            soupdatas = mobile_lists[i].select('a.link_direct')
            temp = []
            for j in soupdatas:
                temp.append([j.text, j.attrs['href']])
            datas.append(temp)
        print(datas)

        mobile_datas = {
            'mobile': datas
        }
        return mobile_datas


if __name__ == '__main__':
    url_pc = "https://www.daum.net/"
    url_mobile = "https://m.daum.net/"

    d1 = get_data(url_pc)
    pcdata = parse_data(d1, 'pc')

    d2 = get_data(url_mobile)
    mobiledata = parse_data(d2, 'mobile')

    pcdata.update(mobiledata)
    write_excel(pcdata)